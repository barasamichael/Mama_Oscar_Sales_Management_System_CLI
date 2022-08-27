#author : Barasa Michael Murunga
#school : Alliance High School
#class : 3Q
#year of development : Wednesday June,2020 15:49pm
#project : MOSMS
#password : mcforijjrikffirejfiinnkfftkjffrfnnrj|
#username : 1jed,e2ikkoo3rsskkmrMkfdo!ffmr,kjFo.rfmo,k|
#key : 8
import sqlite3
import os
import pandas as pd
from sqlalchemy import create_engine
from time import strftime
import openpyxl
from openpyxl import load_workbook
from pandas import read_csv
import pyinputplus as pyip

name = "Mama Oscar Sales Management System"
sep = ('-')*97
month = strftime('%B %G')
date = strftime('%x')
year = strftime('%G')

def run_out_stock():
    print("===========DEPLETED STOCK============")
    data = read_csv(month + '.txt',names = ['product','estimated cost...'])
    print(data)
    diversion()

def update_stock(change,item):
    cursor,conn = Database()
    try:
        cursor.execute("UPDATE `stock " + year +"` SET Quantity = ? WHERE `Product` LIKE ?", (change,'%'+str(item)+'%'))
        conn.commit()
    except:
        print("\n~~~~~ Warning : ~~~~~\n An error occured while updating the stock!!!\n")
    try:
        cursor.execute("SELECT Product, Buying_Price FROM `stock " + year +"` WHERE `Quantity` = 0")
        fetch = cursor.fetchall()
        list1 = []
        for i in fetch:
            list1 = list(i)
        with open(month + ".txt",'a') as file:
            file.write(list1[0]+','+list1[1]+'\n')
    except IndexError:
        pass
    cursor.execute("DELETE FROM `stock " + year +"` WHERE `Quantity` = 0")
    conn.commit()
    cursor.close()
    conn.close()

def view_sales(x):
    data = pd.read_excel(year + ".xlsx",month)
    print(data)
    if x == 1:
        diversion()
    elif x == 0:
        diversion1()

def add_data():
    times = pyip.inputInt(prompt = "How many records do you wish to make? ")
    for i in range(times):
        filename = excel()
        wb = openpyxl.load_workbook(filename)
        sheet = wb[month]
        if (sheet['A1'].value == 'Date') and (sheet['B1'].value == 'Time') and (sheet['C1'].value == 'Item') and (sheet['D1'].value == 'Quantity') and (sheet['E1'].value == 'Cost Price') and (sheet['F1'].value == 'Total'):
           pass 
        else:
           sheet['A1'] = 'Date'
           sheet['B1'] = 'Time'
           sheet['C1'] = 'Item'
           sheet['D1'] = 'Quantity'
           sheet['E1'] = 'Cost Price'
           sheet['F1'] = 'Total'
        first_column = sheet['A']
        second_column = sheet['B']
        third_column = sheet['C']
        fourth_column = sheet['D']
        fifth_column = sheet['E']
        sixth_column = sheet['F']
        col_len1 = str(len(first_column)+1)
        col_len2 = str(len(second_column)+1)
        col_len3 = str(len(third_column)+1)
        col_len4 = str(len(fourth_column)+1)
        col_len5 = str(len(fifth_column)+1)
        col_len6 = str(len(sixth_column)+1)
        print('Provide the data requested below: ')
        item = pyip.inputStr(prompt = "Product : ")
        quantity = pyip.inputInt(prompt = "Quantity : ")
        cursor,conn = Database()
        try:
            cursor.execute("SELECT `Product` FROM `Stock " + year +"` WHERE `Product` LIKE ?", ('%'+str(item)+'%',))
            fetch = cursor.fetchall()
            for i in fetch:
                iproduct = i[0]  
            cursor.execute("SELECT `Quantity` FROM `Stock " + year +"` WHERE `Product` LIKE ?", ('%'+str(item)+'%',))
            fetch = cursor.fetchall()
            for i in fetch:
                amount = i[0]
            cursor.execute("SELECT `Selling_Price` FROM `Stock " + year +"` WHERE `Product` LIKE ?", ('%'+str(item)+'%',))
            fetch = cursor.fetchall()
            for i in fetch:
                price = i[0]
            total = 0
            for i in range(quantity):
                total += int(price)
            sheet['A' + col_len1] = date
            sheet['B' + col_len2] = strftime("%H:%M:%S %p")
            sheet['C'+ col_len3] =  iproduct
            sheet['D' + col_len4] = quantity
            sheet['E' + col_len5] =  price
            sheet['F' + col_len6] =  total
            wb.save(filename)
            change = int(amount) - quantity
            update_stock(change,iproduct)
        except:
            print("\n~~~~~ Warning : ~~~~~\n The requested product was not found in the database!!!\n")
    diversion()

def excel():
    filename = year + ".xlsx"
    if os.path.isfile(filename):
       wb = openpyxl.Workbook(filename)
       if month in wb.get_sheet_names():
           pass
       else:
           wb.create_sheet(index=-1, title = month)
    else:
       wb = openpyxl.Workbook()
       wb.create_sheet(index=-1,title=month)
       wb.save(filename) 
    return filename

def logfile(t,x):
    with open('MOSMS.log','a') as file:
        if t == 1:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : [Program initiated]'+ ';\n')
        if t == 2:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Access level '+ x+ ';\n')
        if t == 7:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Viewed '+ x+ ';\n')
        if t == 3:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Updated item in ' + x + ';\n')
        if t == 4:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Deleted item in ' + x + ';\n')
        if t == 5:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : [Program exit];\n')
        if t == 6:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Access attempt in ' + x + 'failed;\n')
        if t == 8:
            file.write(date +'   '+strftime("%H:%M:%S %p")+' : Searched for ' + x + 'in database;\n')

def timer():
    import time
    print("The attempt you just made is an act of sabotage to the system...unless you do what you think is right, the system is going into lockdown in...")
    def countdown(t):
        while t:
            mins, secs = divmod(t, 60)
            timeformat = '{:02d}:{:02d}'.format(mins, secs)
            print(timeformat, end='\r')
            time.sleep(1)
            t -= 1
        print('*******SYSTEM LOCKDOWN!!!*******')
    countdown(60)
    logfile(5, None)
    exit()

def upper_admin(x):
    print("Enter username and password :")
    username = pyip.inputPassword(prompt = "Username: ",mask = '*')
    password = pyip.inputPassword(prompt = "Password: ",mask = "*")
    admin_verify(username,password)
    if x == 1:
        register_administrators()
    elif x == 0:
        view_administrators()
    diversion()

def validity_administration1(password,username):
        cursor,conn = Database()
        cursor.execute("SELECT * FROM `administrators " + year +"` WHERE `Username` = ? AND `Password` = ?", (username, password))
        while cursor.fetchone() is None:
            print("\nAccess Denied!!!\n")
            logfile(6, "Administrator")
            password,username = admin_control()
            validity_administration1(password,username)
        else:
            cursor.execute("SELECT * FROM `administrators " + year +"` WHERE `Username` = ? AND `Password` = ?", (username, password))
            data = cursor.fetchone()
            admin_id= data[1]
            rank = data[8]
            logfile(2, "Administrator; Name :" + admin_id + " Rank : " + rank)
            print("Access granted!!!")
            admin_menu()
        cursor.close()
        conn.close()

def admin_control():
        print("\nYou are required to provide the username and password to be granted access")
        username = pyip.inputPassword(prompt = "Username: ", mask = "*")
        password = pyip.inputPassword(prompt = "Password: ",mask = "*")
        return password, username

def view_administrators():
    engine = create_engine('sqlite:///'+year+'.db')
    data = pd.read_sql("administrators " + year +"",engine)
    print(data)
    logfile(7, "Administrators database")

def client_menu():
        choice = client()
        if choice == '1':
            view_stock(2)
        elif choice == '2':
            search_stock(2)
        elif choice == '3':
            view_sales(0)
        elif choice == '0':
            exit_choice = pyip.inputYesNo(prompt = "Are you sure you want to exit?(Yes/No) ")
            if exit_choice == 'yes':
                logfile(5, None)
                exit()
            else:
                client_menu()

def client():
    contents = """=====Main Menu=====
    1. View stock
    2. Search stock
    3. View sales
    0. Exit
    """
    values = ['1','2','3','0']
    print(contents)
    choice= pyip.inputMenu(values,prompt = "Make a choice from the main menu: ")
    print(sep)
    return choice

def register_administrators():
    no = pyip.inputInt(prompt = "How many administrators do you wish to add? ")
    for i in range(1,no+1):
        print("\nRecord #%i"%(i))
        id,names,nationality,gender,birth,residence,rank,password,username,email, telephone = administrators()
        cursor,conn = Database()
        try:
            cursor.execute("INSERT INTO `administrators " + year +"` (ID_Number, Names,Nationality,Gender,DOB,Residence,Rank,Password,Username, Email_address, Telephone) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (id,names,nationality,gender,birth,residence,rank,password,username,email, telephone))
            conn.commit()
            logfile(3, 'Administrator database')
        except:
             print("\n~~~~~ Warning : ~~~~~\n An error occured while saving the data to the database...Try again later!!!\n")
        cursor.close()
        conn.close()
        
def administrators():
    id = pyip.inputInt(prompt = "ID :")
    names = pyip.inputStr(prompt = "Names :")
    nationality = pyip.inputStr(prompt = "Nationality :")
    gender = pyip.inputMenu(['M','F','T'],prompt = "Gender(M/F/T) :")
    birth = pyip.inputDate(prompt = "D.O.B. :")
    rank = pyip.inputStr(prompt = "Rank :")
    residence = pyip.inputStr(prompt = "Residence :")
    email = pyip.inputEmail(prompt = "Email Address :")
    telephone = pyip.inputNum(prompt = "Telephone :")
    username = pyip.inputStr(prompt = "Username :")
    password = pyip.inputStr("Password :")  
    return id,names,nationality,gender,birth,residence,rank,password,username,email, telephone

def delete_stock():
    print("Deleting entries is based on ID's of the person; you are therefore required to search the name, confirm the id and delete it...\nWARNING : Deviation from the given instructions can prove catastrophous!!!")
    search_stock(0)
    selection = pyip.inputInt(prompt = "Enter ID to Delete Item: ")
    cursor,conn = Database()
    try:
        cursor.execute("DELETE FROM `stock " + year +"` WHERE `ID` = %d" % selection)
        conn.commit()
        print("\nDeleting item...")
        cursor.close()
        conn.close()
        print("\nItem deleted!")
        logfile(4, 'stock database')
    except:
        print("\n~~~~~ Warning : ~~~~~\n An error occured while deleting the selected item...check the validity of the ID to troubleshoot!!!\n")
    diversion()
def search_stock(x):
    print("Enter keyword to search in files...\nEnter -1 to go back to main menu\n")
    item = pyip.inputStr(prompt = "Enter keyword : ")
    if item == '-1':
        print('\n' + sep + '\n')
        admin_menu()
    elif item != "":
        cursor,conn = Database()
        try:
            cursor.execute("SELECT * FROM `stock " + year +"` WHERE `Product` LIKE ?", ('%'+str(item)+'%',))
            fetch = cursor.fetchall()
            message = ''
            for data in fetch:
                message += str(data)
                print(data)
        except:
            pass
        cursor.close()
        conn.close()
        cipher = encryptMessage(8,message)
        logfile(8, cipher)
    if x == 1:
        diversion()
    elif x == 2:
        diversion1()
    else:
        pass

def encryptMessage(key,message):
    ciphertext = [''] * key
    for column in range(key):
            currentIndex = column
            while currentIndex < len(message):
                ciphertext[column] += message[currentIndex]
                currentIndex += key
    return ''.join(ciphertext)

def view_stock(x):
    engine = create_engine('sqlite:///'+year+'.db')
    data = pd.read_sql('stock '+year+'',engine)
    print(data)
    logfile(7, "stock database")
    if x == 1:
        diversion()
    elif x == 2:
        diversion1()

def diversion():
    input("Press the Enter Key to View Menu: ")
    print('\n' + sep + '\n')
    admin_menu()

def diversion1():
    input("Press the Enter Key to View Menu: ")
    print('\n' + sep + '\n')
    client_menu()

def admin_menu():
        choice = administrator()
        if choice == '1':
            add_stock()   
        elif choice == '2':
            view_stock(1)
        elif choice == '3':
            delete_stock()
        elif choice == '4':
            search_stock(1)
        elif choice == '5':
            upper_admin(1)
        elif choice == '6':
            upper_admin(0)
        elif choice == '#Jubilate':
            admin_register()
        elif choice == '7':
            add_data()
        elif choice == '8':
            view_sales(1)
        elif choice == '9':
            run_out_stock()
        elif choice == '0':
            exit_choice = pyip.inputYesNo(prompt = "Are you sure you want to exit?(Yes/No) ")
            if exit_choice == 'yes':
                logfile(5, None)
                exit()
            else:
                admin_menu()

def items():
    product = pyip.inputStr( prompt = "Product : ")
    quantity = pyip.inputInt( prompt = "Quantity : ")
    price = pyip.inputNum( prompt = "Buying Price : ")
    selling_price = pyip.inputNum( prompt = "Selling price : ")
    return product,price,quantity,selling_price

def add_stock():
    no = pyip.inputInt(prompt = "How many records do you need to add? ")
    for i in range(1,no+1):
        print("\nRecord #%i"%(i))
        product,price,quantity,selling_price = items()
        cursor,conn = Database()
        try:
            cursor.execute("INSERT INTO `stock " + year +"` (Time, Date,Product,Buying_Price,Quantity,Selling_Price) VALUES(?, ?, ?, ?, ?, ?)", (strftime("%H:%M:%S %p"),date,product.title(),price,quantity,selling_price))
            cursor.execute("INSERT INTO `stock-original " + month +"` (Time, Date,Product,Buying_Price,Quantity,Selling_Price) VALUES(?, ?, ?, ?, ?, ?)", (strftime("%H:%M:%S %p"),date,product.title(),price,quantity,selling_price))
            conn.commit()
            logfile(3, 'stock database')
        except:
            print("An error occurred while saving the data in the database...attempt once more...")
        cursor.close()
        conn.close()
    diversion()

def Database():
    conn = sqlite3.connect(year + ".db")
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS `stock " + year + "` (Time TEXT, Date TEXT ,Product_ID INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, Product TEXT, Buying_Price TEXT, Selling_Price TEXT, Quantity TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS `stock-original " + month + "` (Time TEXT, Date TEXT ,Product_ID INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL, Product TEXT, Buying_Price TEXT, Selling_Price TEXT, Quantity TEXT)")
    cursor.execute("CREATE TABLE IF NOT EXISTS `administrators " + year + "` (ID_Number INTEGER PRIMARY KEY, Names TEXT, Nationality TEXT, Gender TEXT, DOB TEXT, Email_address TEXT, Telephone TEXT, Residence TEXT, Rank TEXT, Password TEXT, Username TEXT)")
    return cursor,conn

def administrator():
    contents = """=====Main Menu=====
    1) Add stock
    2) View stock
    3) Delete stock
    4) Search stock
    5) Register Administrators
    6) View Administrators
    7) Sales
    8) View sales
    9) View depleted stock
    0) Quit"""
    print(contents)
    values =  ['1','2','3','4','5','6','#Jubilate','7','8','9','0']
    choice= pyip.inputMenu(values,prompt = "Make a choice from the main menu: ")
    print(sep)
    return choice    

def validate_password():
    """This function allows us to ensure that the password is of required strength and is encoded to protect it."""
    upper_list = "ABCDEFGHIJKLMNOPQRSTUVWXYZ" #provides the uppercase characters
    def contains_upper(s):
        return any(c in ascii_uppercase for c in s)
    from string import ascii_uppercase, ascii_lowercase, digits
    
    def contains(required_chars, s):
        return any(c in required_chars for c in s)
    
    def contains_upper(s):
        return contains(ascii_uppercase, s)
    
    def contains_lower(s):
        return contains(ascii_lowercase, s)
    
    def contains_digit(s):
        return contains(digits, s)
    
    def contains_special(s):
        return contains(r"""!@$%^&*()_-+={}[]|\,.></?~`"':;""", s)
    
    def long_enough(s):
        return len(s) >= 8
    def validate_password(password):
        VALIDATIONS = (
            (contains_upper, 'Password needs at least one upper-case character.'),
            (contains_lower, 'Password needs at least one lower-case character.'),
            (contains_digit, 'Password needs at least one number.'),
            (contains_special, 'Password needs at least one special character.'),
            (long_enough, 'Password needs to be at least 8 characters in length.'),
        )
        failures = [
            msg for validator, msg in VALIDATIONS if not validator(password)
        ]
        if not failures:
            return True
        else:
            print("Invalid password! Review below and change your password accordingly!\n")
            for msg in failures:
                print(msg)
            print('')
            return False
    if __name__ == '__main__':
        while True:
            password = input("Enter desired password: ")
            if validate_password(password):
                break
    return password                

def admin_register():
    print("You are required to enter the password and username.\n")
    username = pyip.inputPassword(prompt = "Username :",mask = '*')
    password = validate_password()
    with open(username + '.txt','w') as file:
        file.write(username + '\n' + password)
    print("Saving...\n\nRegistration successful!!!\n")
    logfile(3, 'Senior Administrator')
    diversion()

def admin_verify(username,password):
    if username + '.txt' in os.listdir():
        with open(username + '.txt',"r") as file:
            verify = file.read().splitlines()
        if password in verify:
            print("\nAccess Granted!!\n")
            logfile(2, "Senior Administrator")
        else :
            print("Password is not recognised!!!")
            logfile(6, "Senior Administrator")
            timer()
    else :
        print("User not found!!!")
        logfile(6, "Senior Administrator")
        timer()     

def main():
    values = ['1','2','0']
    choice = pyip.inputMenu(values,prompt = "Select access level(use the number prior to select the level):\n1. Administrator\n2. Client \n0. Exit\n\n>>>")
    print(sep)
    if choice == '1':
        password,username = admin_control()
        validity_administration1(password,username)
    elif choice == '2':
        logfile(2, "Client")
        client_menu()
    elif choice == '0':
            exit_choice = pyip.inputYesNo(prompt = "Are you sure you want to exit?(Yes/No) ")
            if exit_choice == 'yes':
                logfile(5, None)
                exit()
            else:
                main()

if __name__ == '__main__':
    logfile(1, None)
    print("=======Welcome to Mama Oscar Sales Management System=======")
    main()